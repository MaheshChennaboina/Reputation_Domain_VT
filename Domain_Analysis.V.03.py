import requests
import pandas as pd
import json
import os
import base64
from datetime import datetime
import time
from colorama import init, Fore, Style
import logging
import re
from openpyxl import load_workbook
from openpyxl.styles import Alignment

# Initialize colorama
init()

# Configure logging
logging.basicConfig(filename='url_analysis.log', level=logging.INFO, format='%(asctime)s - %(levelname)s - %(message)s')

# Load API keys from api_keys.json
with open('api_keys.json') as config_file:
    config = json.load(config_file)
api_keys = config['api_keys']
api_key = api_keys[0]  # Use the first API key

# Function to convert URL from hxxp://abc[.]com to http://abc.com
def convert_url(url):
    return url.replace('hxxp', 'http').replace('[.]', '.')

# Function to check the URL without re-analysis
def check_url(url):
    headers = {
        "x-apikey": api_key
    }
    converted_url = convert_url(url)
    encoded_url = base64.urlsafe_b64encode(converted_url.encode()).decode().strip('=')
    response = requests.get(f"https://www.virustotal.com/api/v3/urls/{encoded_url}", headers=headers)
    if response.status_code == 200:
        data = response.json()
        logging.info(f"URL checked successfully: {url}")
        print(Fore.GREEN + f"URL checked successfully: {url}" + Style.RESET_ALL)
        return data['data']['attributes']['last_analysis_stats']['malicious']
    else:
        logging.error(f"Error checking URL: {url}. Response: {response.content}")
        print(Fore.RED + f"Error checking URL: {url}. Response: {response.content}" + Style.RESET_ALL)
        return None

# Function to request a re-analysis of the URL
def request_reanalysis(url):
    headers = {
        "x-apikey": api_key
    }
    converted_url = convert_url(url)
    encoded_url = base64.urlsafe_b64encode(converted_url.encode()).decode().strip('=')
    response = requests.post(f"https://www.virustotal.com/api/v3/urls/{encoded_url}/analyse", headers=headers)
    if response.status_code == 200:
        analysis_id = response.json()['data']['id']
        logging.info(f"Re-analysis requested successfully: {url}")
        print(Fore.GREEN + f"Re-analysis requested successfully: {url}" + Style.RESET_ALL)
        return analysis_id
    else:
        logging.error(f"Error requesting re-analysis: {url}. Response: {response.content}")
        print(Fore.RED + f"Error requesting re-analysis: {url}. Response: {response.content}" + Style.RESET_ALL)
        return None

# Function to get the number of vendors that flagged the URL as malicious after re-analysis
def get_malicious_score(analysis_id, url):
    headers = {
        "x-apikey": api_key
    }
    while True:
        response = requests.get(f"https://www.virustotal.com/api/v3/analyses/{analysis_id}", headers=headers)
        if response.status_code == 200:
            data = response.json()
            status = data['data']['attributes']['status']
            if status == 'completed':
                logging.info(f"Re-analysis completed successfully: {url}")
                print(Fore.GREEN + f"Re-analysis completed successfully: {url}" + Style.RESET_ALL)
                return data['data']['attributes']['stats']['malicious']
            else:
                logging.info(f"Waiting for re-analysis to complete: {url}")
                print(Fore.YELLOW + f"Waiting for re-analysis to complete: {url}" + Style.RESET_ALL)
                time.sleep(15)  # Wait for 15 seconds before checking again
        else:
            logging.error(f"Error retrieving re-analysis results: {url}. Response: {response.content}")
            print(Fore.RED + f"Error retrieving re-analysis results: {url}. Response: {response.content}" + Style.RESET_ALL)
            return None

# Function to get the malicious score based on user selection
def get_malicious_score_with_rate_limit(url, option, index, total):
    if option == '1':
        score = check_url(url)
    elif option == '2':
        analysis_id = request_reanalysis(url)
        if analysis_id:
            score = get_malicious_score(analysis_id, url)
        else:
            score = None
    else:
        score = None
    print(Fore.CYAN + f"Completed {index + 1}/{total}. {total - (index + 1)} left." + Style.RESET_ALL)
    time.sleep(15)  # Rate limit: 4 requests per minute (15 seconds interval)
    return score

# Function to extract IP address from URL
def extract_ip(url):
    ip_pattern = re.compile(r'http://(\d+\.\d+\.\d+\.\d+)')
    match = ip_pattern.search(url)
    if match:
        return match.group(1)
    return None

# Function to get ISP and country information for an IP address
def get_ip_info(ip):
    response = requests.get(f"https://ipinfo.io/{ip}/json")
    if response.status_code == 200:
        data = response.json()
        isp = data.get('org', 'Unknown')
        country = data.get('country', 'Unknown')
        return isp, country
    else:
        logging.error(f"Error retrieving IP info: {ip}. Response: {response.content}")
        print(Fore.RED + f"Error retrieving IP info: {ip}. Response: {response.content}" + Style.RESET_ALL)
        return 'Unknown', 'Unknown'

# Prompt user to select an option
print("Select an option:")
print("1. Check URL")
print("2. Re-analyze URL")
option = input("Enter the option number (1 or 2): ")

# Load the input Excel file
input_file = 'URL_input.xlsx'
df = pd.read_excel(input_file)

# Add new columns for IP addresses, ISP, and country
df['IP Address'] = df['URL'].apply(lambda url: extract_ip(convert_url(url)))
df[['ISP', 'Country']] = df['IP Address'].apply(lambda ip: pd.Series(get_ip_info(ip)) if ip else pd.Series(['Unknown', 'Unknown']))

# Add a new column for the malicious score with rate limiting
total_urls = len(df)
df['Number Vendors Flagged as Malicious(URL)'] = [get_malicious_score_with_rate_limit(url, option, index, total_urls) for index, url in enumerate(df['URL'])]

# Create output directory if it doesn't exist
output_dir = 'URL_response'
if not os.path.exists(output_dir):
    os.makedirs(output_dir)

# Create separate DataFrames for clean, suspicious, and malicious URLs
clean_df = df[df['Number Vendors Flagged as Malicious(URL)'] == 0]
suspicious_df = df[(df['Number Vendors Flagged as Malicious(URL)'] > 0) & (df['Number Vendors Flagged as Malicious(URL)'] < 5)]
malicious_df = df[df['Number Vendors Flagged as Malicious(URL)'] >= 5]

# Get current date and time for file naming
current_datetime = datetime.now().strftime("%Y-%m-%d_%H-%M-%S")

# Save the output to a new Excel file with separate sheets and date in the filename
output_file = os.path.join(output_dir, f'URL_analysis_{current_datetime}.xlsx')
with pd.ExcelWriter(output_file, engine='openpyxl') as writer:
    clean_df.to_excel(writer, sheet_name='Clean', index=False)
    suspicious_df.to_excel(writer, sheet_name='Suspicious', index=False)
    malicious_df.to_excel(writer, sheet_name='Malicious', index=False)

# Open the workbook and adjust column widths for better visibility
wb = load_workbook(output_file)
for sheet_name in ['Clean', 'Suspicious', 'Malicious']:
    ws = wb[sheet_name]
    for row in ws.iter_rows():
        for cell in row:
            cell.alignment = Alignment(horizontal='center', vertical='center')
    for column in ws.columns:
        max_length = 0
        column = [cell for cell in column]
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(cell.value)
            except:
                pass
        adjusted_width = (max_length + 2)
        ws.column_dimensions[column[0].column_letter].width = adjusted_width

wb.save(output_file)

logging.info(f"Output saved to {output_file}")
print(Fore.GREEN + f"Output saved to {output_file}" + Style.RESET_ALL)



import requests
import pandas as pd
import json
import os
import base64
from datetime import datetime
import time
from colorama import init, Fore, Style
import logging
import re
from openpyxl import load_workbook
from openpyxl.styles import Alignment

# Initialize colorama
init()

# Configure logging
logging.basicConfig(filename='url_analysis.log', level=logging.INFO, format='%(asctime)s - %(levelname)s - %(message)s')

# Load API keys from api_keys.json
with open('api_keys.json') as config_file:
    config = json.load(config_file)
api_keys = config['api_keys']
api_key = api_keys[0]  # Use the first API key

# Function to convert URL from hxxp://abc[.]com to http://abc.com
def convert_url(url):
    return url.replace('hxxp', 'http').replace('[.]', '.')

# Function to check the URL without re-analysis
def check_url(url):
    headers = {
        "x-apikey": api_key
    }
    converted_url = convert_url(url)
    encoded_url = base64.urlsafe_b64encode(converted_url.encode()).decode().strip('=')
    response = requests.get(f"https://www.virustotal.com/api/v3/urls/{encoded_url}", headers=headers)
    if response.status_code == 200:
        data = response.json()
        logging.info(f"URL checked successfully: {url}")
        print(Fore.GREEN + f"URL checked successfully: {url}" + Style.RESET_ALL)
        return data['data']['attributes']['last_analysis_stats']['malicious']
    else:
        logging.error(f"Error checking URL: {url}. Response: {response.content}")
        print(Fore.RED + f"Error checking URL: {url}. Response: {response.content}" + Style.RESET_ALL)
        return None

# Function to request a re-analysis of the URL
def request_reanalysis(url):
    headers = {
        "x-apikey": api_key
    }
    converted_url = convert_url(url)
    encoded_url = base64.urlsafe_b64encode(converted_url.encode()).decode().strip('=')
    response = requests.post(f"https://www.virustotal.com/api/v3/urls/{encoded_url}/analyse", headers=headers)
    if response.status_code == 200:
        analysis_id = response.json()['data']['id']
        logging.info(f"Re-analysis requested successfully: {url}")
        print(Fore.GREEN + f"Re-analysis requested successfully: {url}" + Style.RESET_ALL)
        return analysis_id
    else:
        logging.error(f"Error requesting re-analysis: {url}. Response: {response.content}")
        print(Fore.RED + f"Error requesting re-analysis: {url}. Response: {response.content}" + Style.RESET_ALL)
        return None

# Function to get the number of vendors that flagged the URL as malicious after re-analysis
def get_malicious_score(analysis_id, url):
    headers = {
        "x-apikey": api_key
    }
    while True:
        response = requests.get(f"https://www.virustotal.com/api/v3/analyses/{analysis_id}", headers=headers)
        if response.status_code == 200:
            data = response.json()
            status = data['data']['attributes']['status']
            if status == 'completed':
                logging.info(f"Re-analysis completed successfully: {url}")
                print(Fore.GREEN + f"Re-analysis completed successfully: {url}" + Style.RESET_ALL)
                return data['data']['attributes']['stats']['malicious']
            else:
                logging.info(f"Waiting for re-analysis to complete: {url}")
                print(Fore.YELLOW + f"Waiting for re-analysis to complete: {url}" + Style.RESET_ALL)
                time.sleep(15)  # Wait for 15 seconds before checking again
        else:
            logging.error(f"Error retrieving re-analysis results: {url}. Response: {response.content}")
            print(Fore.RED + f"Error retrieving re-analysis results: {url}. Response: {response.content}" + Style.RESET_ALL)
            return None

# Function to get the malicious score based on user selection
def get_malicious_score_with_rate_limit(url, option, index, total):
    if option == '1':
        score = check_url(url)
    elif option == '2':
        analysis_id = request_reanalysis(url)
        if analysis_id:
            score = get_malicious_score(analysis_id, url)
        else:
            score = None
    else:
        score = None
    print(Fore.CYAN + f"Completed {index + 1}/{total}. {total - (index + 1)} left." + Style.RESET_ALL)
    time.sleep(15)  # Rate limit: 4 requests per minute (15 seconds interval)
    return score

# Function to extract IP address from URL
def extract_ip(url):
    ip_pattern = re.compile(r'http://(\d+\.\d+\.\d+\.\d+)')
    match = ip_pattern.search(url)
    if match:
        return match.group(1)
    return None

# Function to get ISP and country information for an IP address
def get_ip_info(ip):
    response = requests.get(f"https://ipinfo.io/{ip}/json")
    if response.status_code == 200:
        data = response.json()
        isp = data.get('org', 'Unknown')
        country = data.get('country', 'Unknown')
        return isp, country
    else:
        logging.error(f"Error retrieving IP info: {ip}. Response: {response.content}")
        print(Fore.RED + f"Error retrieving IP info: {ip}. Response: {response.content}" + Style.RESET_ALL)
        return 'Unknown', 'Unknown'

# Prompt user to select an option
print("Select an option:")
print("1. Check URL")
print("2. Re-analyze URL")
option = input("Enter the option number (1 or 2): ")

# Load the input Excel file
input_file = 'URL_input.xlsx'
df = pd.read_excel(input_file)

# Add new columns for IP addresses, ISP, and country
df['IP Address'] = df['URL'].apply(lambda url: extract_ip(convert_url(url)))
df[['ISP', 'Country']] = df['IP Address'].apply(lambda ip: pd.Series(get_ip_info(ip)) if ip else pd.Series(['Unknown', 'Unknown']))

# Add a new column for the malicious score with rate limiting
total_urls = len(df)
df['Number Vendors Flagged as Malicious(URL)'] = [get_malicious_score_with_rate_limit(url, option, index, total_urls) for index, url in enumerate(df['URL'])]

# Create output directory if it doesn't exist
output_dir = 'URL_response'
if not os.path.exists(output_dir):
    os.makedirs(output_dir)

# Create separate DataFrames for clean, suspicious, and malicious URLs
clean_df = df[df['Number Vendors Flagged as Malicious(URL)'] == 0]
suspicious_df = df[(df['Number Vendors Flagged as Malicious(URL)'] > 0) & (df['Number Vendors Flagged as Malicious(URL)'] < 5)]
malicious_df = df[df['Number Vendors Flagged as Malicious(URL)'] >= 5]

# Get current date and time for file naming
current_datetime = datetime.now().strftime("%Y-%m-%d_%H-%M-%S")

# Save the output to a new Excel file with separate sheets and date in the filename
output_file = os.path.join(output_dir, f'URL_analysis_{current_datetime}.xlsx')
with pd.ExcelWriter(output_file, engine='openpyxl') as writer:
    clean_df.to_excel(writer, sheet_name='Clean', index=False)
    suspicious_df.to_excel(writer, sheet_name='Suspicious', index=False)
    malicious_df.to_excel(writer, sheet_name='Malicious', index=False)

# Open the workbook and adjust column widths for better visibility
wb = load_workbook(output_file)
for sheet_name in ['Clean', 'Suspicious', 'Malicious']:
    ws = wb[sheet_name]
    for row in ws.iter_rows():
        for cell in row:
            cell.alignment = Alignment(horizontal='center', vertical='center')
    for column in ws.columns:
        max_length = 0
        column = [cell for cell in column]
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(cell.value)
            except:
                pass
        adjusted_width = (max_length + 2)
        ws.column_dimensions[column[0].column_letter].width = adjusted_width

wb.save(output_file)

logging.info(f"Output saved to {output_file}")
print(Fore.GREEN + f"Output saved to {output_file}" + Style.RESET_ALL)

