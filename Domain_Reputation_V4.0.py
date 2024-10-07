import os
import requests
import pandas as pd
from datetime import datetime
import json
import socket
import tldextract
import openpyxl
from openpyxl.styles import Font
from colorama import init, Fore, Style


# Initialize Colorama
init(autoreset=True)

def update_dictionary(domain_dict, malicious_vendors, domain):
    domain_dict[domain] = malicious_vendors
    return domain_dict

def load_api_keys(file_path):
    with open(file_path) as f:
        api_keys = json.load(f)
    return api_keys.get('api_keys', [])

def get_domain_details(api_key, domain, domain_dict):
    url = f'https://www.virustotal.com/api/v3/domains/{domain}'
    headers = {'x-apikey': api_key}

    try:
        response = requests.get(url, headers=headers)
        result = response.json()

        if response.status_code == 200:
            whois_info = result['data']['attributes']['whois']
            data_strings = ["Create date: ", "Creation Date: "]
            malicious_vendors = result['data']['attributes']['last_analysis_stats']['malicious']
            for data in data_strings:
                create_date_index = whois_info.find(data)
                if create_date_index >= 0:
                    break

            try:
                print(f"{Fore.YELLOW}Domain: {domain}")
                extracted_domain = tldextract.extract(domain)
                apex_domain = extracted_domain.registered_domain
            except Exception as e:
                print(f"{Fore.RED}Error extracting apex domain for {domain}: {e}")
                apex_domain = None

            print(f"{Fore.GREEN}Apex Domain: {apex_domain}")

            # Extract IP address
            try:
                ip_address = socket.gethostbyname(domain)
            except socket.gaierror:
                ip_address = "Not Resolved"

            print(f"{Fore.GREEN}IP Address: {ip_address}")
            domain_dict = update_dictionary(domain_dict, malicious_vendors, domain)

            if create_date_index != -1:
                create_date_start = create_date_index + len("Creation Date: ")
                create_date_end = create_date_start + 10
                create_date = whois_info[create_date_start:create_date_end].split()[0]
                print(f"{Fore.GREEN}Create Date: {create_date}")
                print(f"{Fore.GREEN}Domain: {domain}.....Processing....done!")
                print(f"{Fore.GREEN}Number of Security Vendors Flagged as Malicious: {malicious_vendors}")

                return create_date, malicious_vendors, ip_address, apex_domain, domain_dict
            else:
                print(f"{Fore.RED}Create date not found in the whois information.")
                return None, None, None, None, domain_dict

        else:
            print(f"{Fore.RED}Error: {result.get('verbose_msg')}")
            return None, None, None, None, domain_dict

    except Exception as e:
        print(f"{Fore.RED}An error occurred: {e}")
        return None, None, None, None, domain_dict

def save_to_excel(results, output_folder):
    timestamp = datetime.now().strftime("%d_%m_%Y_%H%M%S")
    output_path = os.path.join(output_folder, f"Domain_Analysis_Report_{timestamp}.xlsx")
    wb = openpyxl.Workbook()

    # Create separate sheets for different malicious counts
    sheets = {
        'Clean': wb.active,
        'Suspicious': wb.create_sheet('Suspicious'),
        'Malicious': wb.create_sheet('Malicious')
    }

    for sheet_name, sheet in sheets.items():
        # Set the title of the active sheet to 'Clean'
        if sheet_name == 'Clean':
            sheet.title = 'Clean'
        
        # Insert today's date as the first column
        sheet.append(['Date', 'Domain', 'Threat Actor', 'IP Address', 'Create_Date', 'Apex_Domain', 'Malicious_Vendors', 'Final_Verdict'])

        # Adjust column width
        sheet.column_dimensions['A'].width = 12  # Date
        sheet.column_dimensions['B'].width = 30  # Domain
        sheet.column_dimensions['C'].width = 20  # Threat Actor
        sheet.column_dimensions['D'].width = 15  # IP Address
        sheet.column_dimensions['E'].width = 12  # Create Date
        sheet.column_dimensions['F'].width = 30  # Apex Domain
        sheet.column_dimensions['G'].width = 20  # Malicious Vendors
        sheet.column_dimensions['H'].width = 15  # Final Verdict

        # Set header font to bold
        header_font = Font(bold=True)
        for cell in sheet.iter_rows(min_row=1, max_row=1):
            for c in cell:
                c.font = header_font

    # Append data to the appropriate sheet based on the final verdict
    for row in results:
        date = row['Date']  # Get the date from the input file
        domain = row['Domain']
        threat_actor = row['Threat Actor']  # Get threat actor from the input file
        create_date = row['Create_Date']
        malicious_vendors = row['Malicious_Vendors']
        apex_domain = row['Apex_Domain']
        ip_address = row['Main_Domain_IP_Address']
        final_verdict = row['Final_Verdict']

        # Convert tuple to a single value
        if isinstance(final_verdict, tuple):
            final_verdict = final_verdict[0]

        sheet = sheets[final_verdict]

        # Convert tuple values to strings
        if isinstance(malicious_vendors, tuple):
            malicious_vendors = str(malicious_vendors)

        # Append data to the worksheet, now including threat actor
        sheet.append([date, domain, threat_actor, ip_address, create_date, apex_domain, malicious_vendors, final_verdict])

    wb.save(output_path)
    print(f"{Fore.GREEN}Results saved to: {output_path}")

def determine_final_verdict(malicious_vendors):
    if malicious_vendors == 0:
        return "Clean"
    elif 0 < malicious_vendors <= 4:
        return "Suspicious"
    elif malicious_vendors > 4:
        return "Malicious"
    else:
        return "Unknown"

if __name__ == "__main__":
    api_keys_file = "api_keys.json"
    excel_file_path = "input.xlsx"
    output_folder = "Response"

    if not os.path.exists(output_folder):
        os.makedirs(output_folder)

    results = []
    api_keys = load_api_keys(api_keys_file)

    df = pd.read_excel(excel_file_path, header=0)  # Start reading from the first row
    domain_dict = {}
    number_of_requests = 0
    number_domains = 0

    for index, row in df.iterrows():
        try:
            date = row['Date']  # Assuming 'Date' column exists
            domain = row['Domain']
            threat_actor = row['Threat actor Name']  # Assuming 'Threat actor Name' column exists
            domain_dict[domain] = -1
            number_domains += 1
            domain_results = []  # Store results for each domain

            for api_key in api_keys:
                create_date, malicious_vendors, ip_address, apex_domain, domain_dict = get_domain_details(api_key, domain, domain_dict)
                number_of_requests += 1
                if create_date is not None and malicious_vendors is not None:
                    main_domain_verdict = determine_final_verdict(malicious_vendors)

                    if apex_domain != domain and apex_domain in domain_dict:
                        apex_malicious_vendors = domain_dict[apex_domain]
                        malicious_vendors = domain_dict[domain]
                        malicious_vendors = (malicious_vendors, apex_malicious_vendors)
                        final_verdict = determine_final_verdict(apex_malicious_vendors)

                    elif apex_domain != domain and apex_domain not in domain_dict:
                        _, apex_malicious_vendors, _, _, domain_dict = get_domain_details(api_key, apex_domain, domain_dict)
                        number_of_requests += 1
                        malicious_vendors = (malicious_vendors, apex_malicious_vendors)
                        final_verdict = determine_final_verdict(apex_malicious_vendors)
                    else:
                        final_verdict = main_domain_verdict

                    # Append the domain results, including 'Date' and 'Threat actor Name'
                    domain_results.append({
                        'Date': date,
                        'Domain': domain,
                        'Threat Actor': threat_actor,  # Add threat actor here
                        'Create_Date': create_date,
                        'Malicious_Vendors': malicious_vendors,
                        'Apex_Domain': apex_domain,
                        'Main_Domain_IP_Address': ip_address,
                        'Final_Verdict': final_verdict
                    })
                    break  # Break the loop once the domain is successfully scanned with an API key

            # Append results for this domain to the main results list
            results.extend(domain_results)
        except Exception as e:
            print(f"Error processing row {index}: {e}")
            continue

    print(f"Number of Domains Scanned...: {number_domains}")
    print(f"Number of requests done!: {number_of_requests}")
    for key, value in domain_dict.items():
        print(f"Domain: {key} - Malicious count: {value}")

    save_to_excel(results, output_folder)
