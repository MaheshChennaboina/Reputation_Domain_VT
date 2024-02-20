import os
import requests
import pandas as pd
from datetime import datetime
import json
import socket
import tldextract
import openpyxl
from openpyxl.styles import PatternFill
from colorama import init, Fore, Style

# Initialize Colorama
init(autoreset=True)

def load_api_keys(file_path):
    with open(file_path) as f:
        api_keys = json.load(f)
    return api_keys.get('api_keys', [])

def get_domain_details(api_key, domain):
    url = f'https://www.virustotal.com/api/v3/domains/{domain}'
    headers = {'x-apikey': api_key}

    try:
        response = requests.get(url, headers=headers)
        result = response.json()
        
        if response.status_code == 200:
            whois_info = result['data']['attributes']['whois']
            data_strings = list(["Create date: ", "Creation Date: "])
            malicious_vendors = result['data']['attributes']['last_analysis_stats']['malicious']
            for data in data_strings:
                create_date_index = whois_info.find(data)
                if create_date_index >= 0:
                    break
            if create_date_index != -1:
                create_date_start = create_date_index + len("Creation Date: ")
                create_date_end = create_date_start + 10
                create_date = whois_info[create_date_start:create_date_end].split()[0]
                print(f"{Fore.GREEN}Create Date: {create_date}")
                print(f"{Fore.GREEN}Domain: {domain}.....Processing....done!")
                print(f"{Fore.GREEN}Number of Security Vendors Flagged as Malicious: {malicious_vendors}")

                # Extract apex domain
                apex_domain = tldextract.extract(domain).registered_domain
                print(f"{Fore.GREEN}Apex Domain: {apex_domain}")

                # Extract IP address
                ip_address = socket.gethostbyname(domain)
                print(f"{Fore.GREEN}IP Address: {ip_address}")

                return create_date, malicious_vendors, ip_address, apex_domain
            else:
                print(f"{Fore.RED}Create date not found in the whois information.")
                return "unknown", malicious_vendors, None, None

        else:
            print(f"{Fore.RED}Error: {result.get('verbose_msg')}")
            return None, None, None, None

    except Exception as e:
        print(f"{Fore.RED}An error occurred: {e}")
        return None, None, None, None

def save_to_excel(results, output_folder):
    timestamp = datetime.now().strftime("%d_%m_%Y_%H%M%S")
    output_path = os.path.join(output_folder, f"Domain_Analysis_Report_{timestamp}.xlsx")
    wb = openpyxl.Workbook()
    ws = wb.active

    # Insert today's date as the first column
    ws.append(['Date', 'Domain', 'IP Address', 'Create_Date', 'Apex_Domain', 'Malicious_Vendors', 'Final_Verdict'])

    green_fill = PatternFill(start_color="00FF00", end_color="00FF00", fill_type="solid")  # Green color fill

    for row in results:
        Date = datetime.now().strftime("%Y-%m-%d")  # Get today's date
        domain = row['Domain']
        create_date = row['Create_Date']
        malicious_vendors = row['Malicious_Vendors']
        apex_domain = row['Apex_Domain']
        ip_address = row['Main_Domain_IP_Address']
        final_verdict = row['Final_Verdict']

        # Check if main domain and apex domain are the same, and malicious vendors is zero
        if apex_domain == domain and malicious_vendors == 0:
            fill_row_with_green = True
        elif apex_domain != domain and malicious_vendors[1] == 0:
            fill_row_with_green = True    
        else:
            fill_row_with_green = False

        # Convert tuple values to strings
        if isinstance(malicious_vendors, tuple):
            malicious_vendors = str(malicious_vendors)
        if isinstance(final_verdict, tuple):
            final_verdict = str(final_verdict)

        # Append data to the worksheet
        ws.append([Date, domain, ip_address, create_date, apex_domain, malicious_vendors, final_verdict])

        # Apply conditional formatting to the entire row if required
        if fill_row_with_green:
            for cell in ws.iter_rows(min_row=ws.max_row, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
                for c in cell:
                    c.fill = green_fill

    wb.save(output_path)
    print(f"{Fore.GREEN}Results saved to: {output_path}")

def determine_final_verdict(malicious_vendors):
    if malicious_vendors == 0:
        return "Clean"
    elif 0 < malicious_vendors <= 4:
        return "Suspicious"
    elif malicious_vendors > 4:
        return "Malicious"
    else:
        return "Unknown"

if __name__ == "__main__":
    api_keys_file = "api_keys.json"
    excel_file_path = "input.xlsx"
    output_folder = "Response"

    if not os.path.exists(output_folder):
        os.makedirs(output_folder)

    results = []
    api_keys = load_api_keys(api_keys_file)

    df = pd.read_excel(excel_file_path, header=None)

    for index, row in df.iterrows():
        domain = row[0]  # Assuming domain names are in the first column
        for api_key in api_keys:
            create_date, malicious_vendors, ip_address, apex_domain = get_domain_details(api_key, domain)
            if create_date is not None and malicious_vendors is not None:
                main_domain_verdict = determine_final_verdict(malicious_vendors)
                if apex_domain != domain:  # Check if the apex domain is different from the main domain
                    _, apex_malicious_vendors, _, _ = get_domain_details(api_key, apex_domain)
                    malicious_vendors = (malicious_vendors, apex_malicious_vendors)  # Store both main and apex domain malicious vendors as a tuple
                    final_verdict = (main_domain_verdict, determine_final_verdict(apex_malicious_vendors))  # Store both main and apex domain final verdicts as a tuple
                else:
                    final_verdict = main_domain_verdict
                results.append({'Domain': domain, 'Create_Date': create_date, 'Malicious_Vendors': malicious_vendors, 'Apex_Domain': apex_domain, 'Main_Domain_IP_Address': ip_address, 'Final_Verdict': final_verdict})
                break  # Break the loop once a valid result is obtained using one API key

    save_to_excel(results, output_folder)
