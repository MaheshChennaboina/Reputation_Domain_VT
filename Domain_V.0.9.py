import os
import requests
import pandas as pd
from datetime import datetime
import json
import socket
import tldextract
import openpyxl
from openpyxl.styles import PatternFill

def load_api_keys(file_path):
    with open(file_path) as f:
        api_keys = json.load(f)
    return api_keys.get('api_keys', [])

def get_domain_details(api_key, domain):
    url = f'https://www.virustotal.com/api/v3/domains/{domain}'
    headers = {'x-apikey': api_key}

    try:
        response = requests.get(url, headers=headers)
        result = response.json()
        
        if response.status_code == 200:
            whois_info = result['data']['attributes']['whois']
            data_strings = list(["Create date: ", "Creation Date: "])
            malicious_vendors = result['data']['attributes']['last_analysis_stats']['malicious']
            for data in data_strings:
                create_date_index = whois_info.find(data)
                if create_date_index >= 0:
                    break
            if create_date_index != -1:
                create_date_start = create_date_index + len("Creation Date: ")
                create_date_end = create_date_start + 10
                create_date = whois_info[create_date_start:create_date_end].split()[0]
                print(f"Create Date: {create_date}")
                print(f"Domain: {domain}.....Processing....done!")

                print(f"Number of Security Vendors Flagged as Malicious: {malicious_vendors}")

                # Extract apex domain
                apex_domain = tldextract.extract(domain).registered_domain
                print(f"Apex Domain: {apex_domain}")

                # Extract IP address
                ip_address = socket.gethostbyname(domain)
                print(f"IP Address: {ip_address}")

                return create_date, malicious_vendors, ip_address, apex_domain
            else:
                print("Create date not found in the whois information.")
                return "unknown", malicious_vendors, None, None

        else:
            print(f"Error: {result.get('verbose_msg')}")
            return None, None, None, None

    except Exception as e:
        print(f"An error occurred: {e}")
        return None, None, None, None

def save_to_excel(results, output_folder):
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    output_path = os.path.join(output_folder, f"result_{timestamp}.xlsx")
    df = pd.DataFrame(results, columns=['Domain', 'Create_Date', 'Malicious_Vendors', 'Apex_Domain', 'Main_Domain_IP_Address', 'Final_Verdict'])
    df.to_excel(output_path, index=False)

    # Apply conditional formatting
    wb = openpyxl.load_workbook(output_path)
    ws = wb.active
    green_fill = PatternFill(start_color="00FF00", end_color="00FF00", fill_type="solid")  # Green color fill

    for row_idx, row in enumerate(ws.iter_rows(min_row=2, max_row=len(results) + 1, min_col=3, max_col=3), start=2):
        if row[0].value == 0:
            for cell in ws[row_idx]:
                cell.fill = green_fill  # Apply fill to all cells in the row

    wb.save(output_path)
    print(f"Results saved to: {output_path}")

def determine_final_verdict(malicious_vendors):
    if malicious_vendors == 0:
        return "Clean"
    elif 0 < malicious_vendors <= 4:
        return "Suspicious"
    elif malicious_vendors > 4:
        return "Malicious"
    else:
        return "Unknown"

if __name__ == "__main__":
    api_keys_file = "api_keys.json"
    excel_file_path = "input.xlsx"
    output_folder = "Response"

    if not os.path.exists(output_folder):
        os.makedirs(output_folder)

    results = []
    api_keys = load_api_keys(api_keys_file)

    df = pd.read_excel(excel_file_path, header=None)

    for index, row in df.iterrows():
        domain = row[0]  # Assuming domain names are in the first column
        for api_key in api_keys:
            create_date, malicious_vendors, ip_address, apex_domain = get_domain_details(api_key, domain)
            if create_date is not None and malicious_vendors is not None:
                main_domain_verdict = determine_final_verdict(malicious_vendors)
                if apex_domain != domain:  # Check if the apex domain is different from the main domain
                    _, apex_malicious_vendors, _, _ = get_domain_details(api_key, apex_domain)
                    malicious_vendors = (malicious_vendors, apex_malicious_vendors)  # Store both main and apex domain malicious vendors as a tuple
                    final_verdict = (main_domain_verdict, determine_final_verdict(apex_malicious_vendors))  # Store both main and apex domain final verdicts as a tuple
                else:
                    final_verdict = main_domain_verdict
                results.append({'Domain': domain, 'Create_Date': create_date, 'Malicious_Vendors': malicious_vendors, 'Apex_Domain': apex_domain, 'Main_Domain_IP_Address': ip_address, 'Final_Verdict': final_verdict})
                break  # Break the loop once a valid result is obtained using one API key

    save_to_excel(results, output_folder)
